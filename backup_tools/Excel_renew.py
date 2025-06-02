# -*- coding: utf-8 -*-

# ==============================================================================
# 脚本功能核心备注 (Script Core Functionality Notes)
# ==============================================================================
#
# 脚本名称 (Script Name):
#   Excel_renew.py
#
# 主要目的 (Main Purpose):
#   本脚本用于批量处理指定文件夹内的Excel文件，对每个文件执行以下操作：
#   1. 清空所有工作表的K2单元格内容
#   2. 清空第一个工作表的C列内容（保留C1单元格）
#
# 工作流程 (Workflow):
#   1. 提示用户输入包含Excel文件的文件夹路径
#   2. 验证用户输入的路径是否为有效文件夹
#   3. 扫描文件夹中的所有.xlsx文件（排除临时文件）
#   4. 对每个Excel文件执行清空操作
#   5. 显示详细的处理进度和统计信息
#   6. 生成最终的处理报告
#
# 优化特性 (Optimization Features):
#   - 添加了类型提示，提高代码可读性和IDE支持
#   - 改进了错误处理机制，提供更详细的错误信息
#   - 增强了用户体验，包括进度显示和配置确认
#   - 添加了详细的统计信息和性能监控
#   - 支持用户中断操作（Ctrl+C）优雅退出
#   - 改进了文件验证和安全性检查
#
# 注意事项 (Important Notes):
#   - 此操作会直接修改原始Excel文件，请确保在操作前备份重要数据
#   - 脚本需要安装openpyxl库：pip install openpyxl
#   - 如果Excel文件被其他程序打开，可能会导致保存失败
#   - 支持用户中断操作（Ctrl+C）优雅退出
#
# ==============================================================================

import os
import time
import sys
from pathlib import Path

# Python 3.7兼容的类型提示导入
try:
    from typing import Tuple, Optional
except ImportError:
    # 如果typing模块导入失败，定义空的类型提示
    Tuple = tuple
    Optional = type(None)

try:
    import openpyxl
except ImportError:
    print("错误：缺少 'openpyxl' 库。请先使用 'pip install openpyxl' 命令安装。")
    sys.exit(1)


def get_valid_folder_path_from_user(prompt_message: str) -> Path:
    """
    获取用户输入的有效文件夹路径。
    
    参数:
        prompt_message (str): 提示用户输入的消息。
    
    返回:
        Path: 有效的文件夹路径对象。
    """
    while True:
        try:
            user_input = input(prompt_message).strip()
            
            if not user_input:
                print("错误：未输入路径，请重新输入。")
                continue
                
            # 处理引号包围的路径
            if user_input.startswith('"') and user_input.endswith('"'):
                user_input = user_input[1:-1]
            elif user_input.startswith("'") and user_input.endswith("'"):
                user_input = user_input[1:-1]
            
            folder_path = Path(user_input)
            
            if not folder_path.exists():
                print(f"错误：路径 '{folder_path}' 不存在。请重新输入。")
                continue
                
            if not folder_path.is_dir():
                print(f"错误：路径 '{folder_path}' 不是一个有效的文件夹。请重新输入。")
                continue
                
            return folder_path
            
        except KeyboardInterrupt:
            print("\n操作已由用户中止。")
            sys.exit(0)
        except Exception as e:
            print(f"错误：处理路径时发生异常: {e}。请重新输入。")


def clear_cells_in_excel(file_path: Path) -> Tuple[bool, int, int]:
    """
    针对单个Excel文件执行清空操作。
    
    参数:
        file_path (Path): Excel文件路径。
    
    返回:
        tuple: (是否成功, 清空的K2单元格数量, 清空的C列单元格数量)
    """
    print(f"\n===== 开始处理文件: {file_path.name} =====")
    
    try:
        # 检查文件是否存在
        if not file_path.is_file():
            print(f"错误：文件 '{file_path}' 不是一个有效的文件。跳过...")
            return False, 0, 0

        # 加载Excel工作簿
        print("正在加载文件...")
        workbook = openpyxl.load_workbook(file_path)
        print("文件加载成功，开始处理工作表...")

        k2_cleared_count = 0
        c_column_cleared_count = 0

        # 步骤1: 清空所有工作表的K2单元格
        print("--- 开始清空所有工作表的K2单元格 ---")
        sheet_names = workbook.sheetnames
        
        for i, sheet_name in enumerate(sheet_names, 1):
            worksheet = workbook[sheet_name]
            try:
                # 直接设置K2单元格为None（openpyxl会自动创建单元格）
                if worksheet['K2'].value is not None:
                    worksheet['K2'].value = None
                    k2_cleared_count += 1
                    print(f"  [{i}/{len(sheet_names)}] 已清空工作表 '{sheet_name}' 的K2单元格")
                else:
                    print(f"  [{i}/{len(sheet_names)}] 工作表 '{sheet_name}' 的K2单元格已为空")
            except Exception as e_k2:
                print(f"  [{i}/{len(sheet_names)}] 处理工作表 '{sheet_name}' 的K2单元格时出现问题: {e_k2}")
        
        print(f"--- K2单元格清空完成，共清空 {k2_cleared_count} 个单元格 ---")

        # 步骤2: 清空第一个工作表的C列（从C2开始）
        if sheet_names:
            first_sheet = workbook.worksheets[0]
            first_sheet_name = first_sheet.title
            print(f"\n--- 开始清空第一个工作表 ('{first_sheet_name}') 的C列 (从第2行开始) ---")

            # 获取工作表的最大行数
            max_row = first_sheet.max_row

            if max_row >= 2:
                # 从第2行开始迭代，直到工作表的最大行
                for row_index in range(2, max_row + 1):
                    cell_ref = f'C{row_index}'
                    if first_sheet[cell_ref].value is not None:
                        first_sheet[cell_ref].value = None
                        c_column_cleared_count += 1
                        
                    # 显示进度（每处理50行显示一次）
                    if row_index % 50 == 0 or row_index == max_row:
                        print(f"  处理进度: {row_index}/{max_row} 行")
                
                print(f"  在工作表 '{first_sheet_name}' 的C列中，从第2行到第{max_row}行，共清空了 {c_column_cleared_count} 个有内容的单元格")
            else:
                print(f"  工作表 '{first_sheet_name}' 的总行数不足2行，无需清空C列")
            
            print(f"--- 第一个工作表 ('{first_sheet_name}') C列处理完成 ---")
        else:
            print("\n警告：工作簿中没有找到任何工作表")

        # 保存修改后的工作簿到原文件
        print("\n正在保存更改...")
        workbook.save(file_path)
        print(f"成功！文件 '{file_path.name}' 已按要求处理并保存")
        print(f"===== 文件处理完毕: {file_path.name} =====")
        
        return True, k2_cleared_count, c_column_cleared_count

    except ImportError:
        print("错误：缺少 'openpyxl' 库。请先使用 'pip install openpyxl' 命令安装。")
        raise
    except Exception as e:
        print(f"\n处理文件 '{file_path.name}' 时发生错误：{e}")
        print("请检查文件是否已关闭、文件是否损坏、以及是否有读写权限。将跳过此文件。")
        print(f"===== 文件处理失败: {file_path.name} =====")
        return False, 0, 0


def scan_excel_files(folder_path: Path) -> list:
    """
    扫描文件夹中的Excel文件。
    
    参数:
        folder_path (Path): 文件夹路径。
    
    返回:
        list: Excel文件路径列表。
    """
    excel_files = []
    
    try:
        for file_path in folder_path.iterdir():
            if file_path.is_file():
                # 检查文件是否是.xlsx文件（忽略大小写）
                if file_path.suffix.lower() == '.xlsx':
                    # 排除Excel临时文件（通常以~$开头）
                    if not file_path.name.startswith('~$'):
                        excel_files.append(file_path)
    except Exception as e:
        print(f"扫描文件夹时发生错误: {e}")
    
    return excel_files


def process_excel_files_batch() -> Tuple[bool, int, int, int, int]:
    """
    批量处理Excel文件的主函数。
    支持两种输入模式：
    1. 交互式输入模式（命令行直接运行）
    2. 标准输入模式（Web环境或管道输入）
    
    返回:
        tuple: (是否全部成功, 成功处理的文件数, 失败的文件数, 总K2清空数, 总C列清空数)
    """
    print("Excel文件批量处理工具 (Claude4优化版)")
    print("=" * 60)
    
    start_time = time.time()
    
    try:
        # 1. 智能检测输入模式并获取文件夹路径
        print("\n步骤 1: 获取文件夹路径")
        
        # 检测是否为非交互模式（Web环境或管道输入）
        is_non_interactive = hasattr(sys.stdin, 'isatty') and not sys.stdin.isatty()
        
        if is_non_interactive:
            # 非交互模式：从标准输入读取参数（适用于Web环境）
            print("🌐 检测到Web环境，使用标准输入模式")
            try:
                path_str = input().strip()
                folder_path = Path(path_str)
                
                if not folder_path.exists() or not folder_path.is_dir():
                    raise ValueError(f"路径不存在或不是目录: {path_str}")
                    
            except (ValueError, EOFError) as e:
                print(f"❌ 参数读取错误: {e}")
                return False, 0, 0, 0, 0
        else:
            # 交互模式：使用原有的交互式输入函数
            print("💻 检测到命令行环境，使用交互式输入模式")
            folder_path = get_valid_folder_path_from_user(
                "请输入包含Excel文件的文件夹路径: "
            )
        
        print(f"\n✅ 配置确认:")
        print(f"- 目标文件夹: {folder_path}")
        print(f"- 处理文件类型: .xlsx文件")
        print(f"- 操作内容: 清空所有工作表的K2单元格 + 清空第一个工作表的C列（从C2开始）")
        
        # 2. 扫描Excel文件
        print("\n步骤 2: 扫描Excel文件")
        excel_files = scan_excel_files(folder_path)
        
        if not excel_files:
            print(f"⚠️ 警告：在文件夹 '{folder_path}' 中没有找到任何.xlsx文件")
            return False, 0, 0, 0, 0
        
        print(f"找到 {len(excel_files)} 个Excel文件:")
        for i, file_path in enumerate(excel_files, 1):
            print(f"  {i}. {file_path.name}")
        
        # 3. 自动开始处理（Web环境下不需要用户确认）
        print(f"\n步骤 3: 开始处理 {len(excel_files)} 个文件")
        print("⚠️ 警告：此操作将直接修改Excel文件内容！")
        print("=" * 60)
        
        processed_files_count = 0
        error_files_count = 0
        total_k2_cleared = 0
        total_c_column_cleared = 0
        failed_files = []
        
        for i, file_path in enumerate(excel_files, 1):
            print(f"\n[{i}/{len(excel_files)}] 处理文件: {file_path.name}")
            
            try:
                success, k2_count, c_count = clear_cells_in_excel(file_path)
                
                if success:
                    processed_files_count += 1
                    total_k2_cleared += k2_count
                    total_c_column_cleared += c_count
                    print(f"✅ 成功处理 - K2清空: {k2_count}, C列清空: {c_count}")
                else:
                    error_files_count += 1
                    failed_files.append(file_path.name)
                    print(f"❌ 处理失败")
                    
            except KeyboardInterrupt:
                print("\n\n操作被用户中断")
                break
            except Exception as e:
                error_files_count += 1
                failed_files.append(file_path.name)
                print(f"❌ 处理文件时发生未预期的错误: {e}")
        
        # 5. 生成处理报告
        execution_time = time.time() - start_time
        
        print("\n" + "=" * 60)
        print("📊 处理完成 - 统计报告")
        print("=" * 60)
        print(f"📁 处理文件夹: {folder_path}")
        print(f"📄 扫描到的Excel文件: {len(excel_files)} 个")
        print(f"✅ 成功处理: {processed_files_count} 个文件")
        print(f"❌ 处理失败: {error_files_count} 个文件")
        print(f"🔧 总计清空K2单元格: {total_k2_cleared} 个")
        print(f"🔧 总计清空C列单元格: {total_c_column_cleared} 个")
        print(f"⏱️  总执行时间: {execution_time:.2f} 秒")
        
        if processed_files_count > 0:
            print(f"📈 平均处理速度: {processed_files_count/execution_time:.2f} 文件/秒" if execution_time > 0 else "📈 平均处理速度: N/A")
        
        if failed_files:
            print(f"\n⚠️  处理失败的文件:")
            for failed_file in failed_files:
                print(f"   - {failed_file}")
        
        print("=" * 60)
        
        return error_files_count == 0, processed_files_count, error_files_count, total_k2_cleared, total_c_column_cleared
        
    except KeyboardInterrupt:
        print("\n\n操作被用户中断")
        return False, 0, 0, 0, 0
    except Exception as e:
        print(f"\n批量处理过程中发生错误: {e}")
        return False, 0, 0, 0, 0


def main():
    """
    主函数：协调整个Excel文件批量处理流程。
    """
    try:
        print("🚀 启动Excel文件批量处理工具...")
        
        success, processed_count, error_count, k2_count, c_count = process_excel_files_batch()
        
        if success and processed_count > 0:
            print("\n🎉 所有文件处理成功！")
        elif processed_count > 0:
            print(f"\n⚠️  部分文件处理完成，{error_count} 个文件处理失败")
        else:
            print("\n❌ 没有文件被成功处理")
            
    except KeyboardInterrupt:
        print("\n\n👋 程序被用户中断，感谢使用！")
    except Exception as e:
        print(f"\n💥 程序运行时发生未预期的错误: {e}")
        print("请检查环境配置或联系技术支持")
    finally:
        print("\n程序结束")


if __name__ == "__main__":
    main()