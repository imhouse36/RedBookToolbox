# 导入所需的库
import openpyxl
import os

# 定义清空单元格函数
def clear_cells_in_excel(file_path):
    """
    主要功能:
        针对单个 Excel 文件执行以下操作：
        1. 打开指定的 Excel 文件，遍历其中的每一个工作表 (Sheet)，
           并将每个工作表中 K2 单元格的内容清空。
        2. 额外清空第一个工作表的 C 列内容，保留 C1 单元格，
           即从 C2 单元格开始清空直到该列有数据的最后一行。

    工作过程:
        (处理单个文件的详细过程见下方实现)

    达成结果:
        - 该 Excel 文件中，所有工作表的 K2 单元格内容被清除。
        - 该文件的第一个工作表中，C 列从第 2 行开始的所有单元格内容被清除。
        - 文件会被覆盖保存。

    注意事项:
        - 此操作会直接修改原始 Excel 文件，请确保在操作前备份重要数据。
        - 脚本需要安装 openpyxl 库。
        - 如果 Excel 文件被其他程序打开，可能会导致保存失败。
        - 函数本身处理单个文件，由调用者负责遍历文件夹。
    """
    print(f"\n===== 开始处理文件: {os.path.basename(file_path)} =====") # 增加文件名标识
    try:
        # 检查文件是否存在 (虽然调用前已检查，双重保险)
        if not os.path.isfile(file_path):
            print(f"错误：文件 '{file_path}' 不是一个有效的文件。跳过...")
            return # 文件不存在则退出函数

        # 加载 Excel 工作簿
        print(f"正在加载文件...") # 不再重复完整路径
        workbook = openpyxl.load_workbook(file_path)
        print("文件加载成功，开始处理工作表...")

        # --- 步骤 1: 清空所有工作表的 K2 单元格 ---
        print("--- 开始清空所有工作表的 K2 单元格 ---")
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            try:
                # 检查单元格是否存在，如果不存在则不操作 (虽然openpyxl通常会创建)
                if 'K2' in worksheet._cells:
                    worksheet['K2'].value = None
                    print(f"  - 已清空工作表 '{sheet_name}' 的 K2 单元格。")
                else:
                    print(f"  - 工作表 '{sheet_name}' 中不存在 K2 单元格。")
            except Exception as e_k2:
                print(f"  - 处理工作表 '{sheet_name}' 的 K2 单元格时出现问题: {e_k2}")
        print("--- K2 单元格清空完成 ---")

        # --- 步骤 2: 清空第一个工作表的 C 列 (从 C2 开始) ---
        if sheet_names: # 确保至少有一个工作表
            first_sheet = workbook.worksheets[0] # 获取第一个工作表对象
            first_sheet_name = first_sheet.title
            print(f"\n--- 开始清空第一个工作表 ('{first_sheet_name}') 的 C 列 (从第 2 行开始) ---")

            # 获取工作表的最大行数
            max_row = first_sheet.max_row

            # 只有当最大行数 >= 2 时，才需要清空 C2 及之后的单元格
            if max_row >= 2:
                cleared_count = 0
                # 从第 2 行开始迭代，直到工作表的最大行
                for row_index in range(2, max_row + 1):
                    cell_ref = f'C{row_index}'
                    # 检查单元格是否存在值
                    if first_sheet[cell_ref].value is not None:
                        first_sheet[cell_ref].value = None
                        cleared_count += 1
                print(f"  - 在工作表 '{first_sheet_name}' 的 C 列中，从第 2 行到第 {max_row} 行，共清空了 {cleared_count} 个有内容的单元格。")
            else:
                print(f"  - 工作表 '{first_sheet_name}' 的总行数不足 2 行，无需清空 C 列。")
            print(f"--- 第一个工作表 ('{first_sheet_name}') C 列处理完成 ---")
        else:
            print("\n警告：工作簿中没有找到任何工作表。")

        # 保存修改后的工作簿到原文件
        print("\n正在保存更改...")
        workbook.save(file_path)
        print(f"成功！文件 '{os.path.basename(file_path)}' 已按要求处理并保存。")
        print(f"===== 文件处理完毕: {os.path.basename(file_path)} =====") # 结束标识

    except ImportError:
        # 这个错误通常在脚本启动时就会发现，但保留以防万一
        print("错误：缺少 'openpyxl' 库。请先使用 'pip install openpyxl' 命令安装。")
        # 遇到库缺失问题，可能需要停止整个脚本
        raise # 重新抛出异常，让调用者知道出错了
    except Exception as e:
        # 捕获其他可能的错误，例如文件损坏、权限问题等
        print(f"\n处理文件 '{os.path.basename(file_path)}' 时发生错误：{e}")
        print("请检查文件是否已关闭、文件是否损坏、以及是否有读写权限。将跳过此文件。")
        print(f"===== 文件处理失败: {os.path.basename(file_path)} =====") # 失败标识

# 脚本主入口
if __name__ == "__main__":
    # 提示用户输入文件夹路径
    while True:
        folder_path = input("请输入包含 Excel 文件的文件夹目录路径 (例如 D:\\path\\to\\your\\folder): ")
        # 检查输入的是否是一个存在的目录
        if os.path.isdir(folder_path):
            break
        else:
            print("输入无效，请输入一个有效的文件夹目录路径。")

    print(f"\n开始扫描文件夹: {folder_path}")
    processed_files_count = 0
    error_files_count = 0

    # 遍历指定目录下的所有文件和文件夹
    for filename in os.listdir(folder_path):
        # 检查文件是否是以 .xlsx 结尾 (忽略大小写)
        if filename.lower().endswith('.xlsx'):
            # 并且不是 Excel 临时文件 (通常以 ~$ 开头)
            if not filename.startswith('~$'):
                # 构建文件的完整路径
                full_file_path = os.path.join(folder_path, filename)
                # 确认这确实是一个文件 (以防万一 listdir 返回奇怪的东西)
                if os.path.isfile(full_file_path):
                    try:
                        # 调用核心函数处理这个 Excel 文件
                        clear_cells_in_excel(full_file_path)
                        processed_files_count += 1
                    except Exception:
                        # 如果 clear_cells_in_excel 内部捕获并打印了错误,
                        # 或者抛出了未捕获的异常 (比如 ImportError)
                        error_files_count += 1
                        # 这里可以根据需要决定是否因为一个文件的错误停止整个脚本
                        # print("由于发生错误，脚本终止。")
                        # break # 如果需要停止，取消注释这行
                        pass # 当前选择忽略错误，继续处理下一个文件

    # 结束时打印总结信息
    print("\n==================== 处理完成 ====================")
    print(f"总共扫描了文件夹: {folder_path}")
    print(f"成功处理了 {processed_files_count} 个 Excel (.xlsx) 文件。")
    if error_files_count > 0:
        print(f"有 {error_files_count} 个文件在处理过程中发生错误 (详见上方日志)。")
    print("==================================================")