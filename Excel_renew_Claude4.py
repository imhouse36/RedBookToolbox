# -*- coding: utf-8 -*-
# 脚本功能:
# 本脚本用于批量处理指定文件夹内的所有Excel文件（.xlsx格式）。
# 对每个Excel文件执行以下操作：
#   1. 清空所有工作表的K2单元格内容
#   2. 清空第一个工作表的整个C列内容
#   3. 保存文件
#
# 工作流程:
#   1. 提示用户输入要处理的文件夹路径。
#   2. 校验文件夹路径是否存在。
#   3. 扫描文件夹内的所有.xlsx文件。
#   4. 逐个处理每个Excel文件：
#      a. 打开Excel文件
#      b. 遍历所有工作表，清空K2单元格
#      c. 清空第一个工作表的C列
#      d. 保存文件
#   5. 输出处理结果和统计信息。
#
# 达成的结果:
# - 指定文件夹内所有Excel文件的K2单元格将被清空。
# - 每个Excel文件的第一个工作表的C列将被完全清空。
# - 控制台会输出详细的处理状态和最终的统计报告。
#
# 注意事项:
# - 脚本仅处理.xlsx格式的Excel文件，不支持.xls格式。
# - 请确保Excel文件没有被其他程序打开，否则可能导致处理失败。
# - 建议在处理前备份重要文件。
# - 请确保对目标文件夹及其文件有读写权限。

import pathlib
import time
import sys
from typing import List, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError:
    print("错误：缺少必要的依赖库 'openpyxl'")
    print("请运行以下命令安装：pip install openpyxl")
    sys.exit(1)

def get_valid_folder_path_from_user(prompt_message: str) -> pathlib.Path:
    """
    提示用户输入一个文件夹路径，并持续请求直到输入一个已存在的有效文件夹路径。

    参数:
        prompt_message (str): 显示给用户的提示信息。

    返回:
        pathlib.Path: 用户输入的有效文件夹路径对象。
    """
    while True:
        try:
            folder_path_str = input(prompt_message).strip()
            if not folder_path_str:
                print("错误：未输入路径。请重新输入。")
                continue
            
            folder_path = pathlib.Path(folder_path_str)
            
            if folder_path.exists() and folder_path.is_dir():
                return folder_path
            else:
                print(f"错误：路径 '{folder_path}' 不存在或不是一个文件夹。请重新输入。")
        except KeyboardInterrupt:
            print("\n用户取消操作。")
            sys.exit(0)

def get_excel_files_from_folder(folder_path: pathlib.Path) -> List[pathlib.Path]:
    """
    获取指定文件夹中的所有Excel文件（.xlsx格式）。

    参数:
        folder_path (pathlib.Path): 文件夹路径。

    返回:
        List[pathlib.Path]: Excel文件路径列表。
    """
    print(f"\n--- 扫描文件夹 '{folder_path}' ---")
    
    try:
        excel_files = [
            f for f in folder_path.iterdir()
            if f.is_file() and f.suffix.lower() == '.xlsx'
        ]
        
        print(f"找到 {len(excel_files)} 个Excel文件")
        
        if excel_files:
            print("Excel文件列表:")
            for i, excel_file in enumerate(excel_files, 1):
                file_size = excel_file.stat().st_size / 1024  # KB
                print(f"  {i}. {excel_file.name} ({file_size:.1f} KB)")
        
        return excel_files
        
    except OSError as e:
        print(f"错误：无法读取文件夹 '{folder_path}': {e}")
        return []

def process_excel_file(excel_file_path: pathlib.Path) -> bool:
    """
    处理单个Excel文件：清空所有工作表的K2单元格和第一个工作表的C列。

    参数:
        excel_file_path (pathlib.Path): Excel文件路径。

    返回:
        bool: 处理是否成功。
    """
    try:
        # 加载工作簿
        workbook = load_workbook(excel_file_path)
        
        # 获取所有工作表名称
        worksheet_names = workbook.sheetnames
        
        # 处理所有工作表的K2单元格
        k2_cleared_count = 0
        for sheet_name in worksheet_names:
            worksheet = workbook[sheet_name]
            if worksheet['K2'].value is not None:
                worksheet['K2'].value = None
                k2_cleared_count += 1
        
        # 处理第一个工作表的C列
        c_column_cleared = False
        if worksheet_names:
            first_worksheet = workbook[worksheet_names[0]]
            
            # 清空C列的所有单元格
            max_row = first_worksheet.max_row
            if max_row > 0:
                for row in range(1, max_row + 1):
                    if first_worksheet[f'C{row}'].value is not None:
                        first_worksheet[f'C{row}'].value = None
                        c_column_cleared = True
        
        # 保存文件
        workbook.save(excel_file_path)
        workbook.close()
        
        # 输出处理详情
        details = []
        if k2_cleared_count > 0:
            details.append(f"K2单元格清空: {k2_cleared_count}个工作表")
        if c_column_cleared:
            details.append("C列已清空")
        
        if details:
            print(f"    处理详情: {', '.join(details)}")
        else:
            print(f"    无需处理（K2和C列均为空）")
        
        return True
        
    except PermissionError:
        print(f"    错误: 文件被占用或权限不足")
        return False
    except Exception as e:
        print(f"    错误: {e}")
        return False

def process_excel_files(excel_files: List[pathlib.Path]) -> Tuple[int, int]:
    """
    批量处理Excel文件。

    参数:
        excel_files (List[pathlib.Path]): Excel文件路径列表。

    返回:
        Tuple[int, int]: (成功处理的文件数量, 失败的文件数量)
    """
    print(f"\n--- 开始批量处理Excel文件 ---")
    
    success_count = 0
    failed_count = 0
    total_files = len(excel_files)
    
    for i, excel_file in enumerate(excel_files, 1):
        progress = (i / total_files) * 100
        print(f"  [进度: {progress:.1f}%] 处理文件: {excel_file.name}")
        
        if process_excel_file(excel_file):
            success_count += 1
            print(f"    ✓ 处理成功")
        else:
            failed_count += 1
            print(f"    ✗ 处理失败")
    
    return success_count, failed_count

def main():
    """
    主函数：控制程序的执行流程。
    """
    print("Excel文件批量处理工具")
    print("功能：清空K2单元格和第一个工作表的C列")
    print("=" * 50)
    
    # 记录脚本开始时间
    start_time = time.time()
    
    try:
        # 1. 获取用户输入：文件夹路径
        folder_path = get_valid_folder_path_from_user(
            "请输入包含Excel文件的文件夹路径: "
        )
        
        # 2. 获取文件夹中的Excel文件
        excel_files = get_excel_files_from_folder(folder_path)
        
        if not excel_files:
            print("\n未找到任何Excel文件（.xlsx格式）。")
            print("请确认文件夹中包含.xlsx格式的Excel文件。")
            return
        
        print(f"\n配置确认:")
        print(f"- 目标文件夹: {folder_path}")
        print(f"- Excel文件数量: {len(excel_files)}")
        print(f"- 处理操作: 清空所有工作表的K2单元格 + 清空第一个工作表的C列")
        
        # 确认处理
        confirm = input("\n确认开始处理？(y/N): ").strip().lower()
        if confirm not in ['y', 'yes', '是']:
            print("用户取消操作。")
            return
        
        # 3. 执行批量处理
        success_count, failed_count = process_excel_files(excel_files)
        
        # 4. 输出统计信息
        end_time = time.time()
        execution_time = end_time - start_time
        
        print(f"\n{'='*60}")
        print("Excel文件处理完成统计报告:")
        print(f"{'='*60}")
        print(f"成功处理文件数量: {success_count}")
        print(f"处理失败文件数量: {failed_count}")
        print(f"总文件数量: {len(excel_files)}")
        print(f"总执行时间: {execution_time:.2f} 秒")
        
        if success_count > 0:
            print(f"平均处理速度: {success_count/execution_time:.2f} 文件/秒")
        
        success_rate = (success_count / len(excel_files)) * 100 if excel_files else 0
        print(f"成功率: {success_rate:.1f}%")
        
        print(f"{'='*60}")
        
        if failed_count > 0:
            print("\n注意：部分文件处理失败，可能原因：")
            print("- 文件被其他程序打开")
            print("- 文件权限不足")
            print("- 文件损坏或格式异常")
            print("建议检查失败的文件并重试。")
        
        print("\n程序执行完毕。")
        
    except KeyboardInterrupt:
        print("\n\n用户中断程序执行。")
    except Exception as e:
        print(f"\n程序执行过程中发生意外错误: {e}")
        print("请检查输入参数和文件权限后重试。")

if __name__ == "__main__":
    main()